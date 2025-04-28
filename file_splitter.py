from openpyxl import load_workbook
import csv
import os

def split_excel_to_csvs(
    excel_path,
    output_dir,
    max_rows_per_file=50000,  # adjust to taste
    sheet_name=None
):
    os.makedirs(output_dir, exist_ok=True)
    
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    
    # Grab header row (no need to access .value here since `values_only=True` gives you the value directly)
    header = next(ws.iter_rows(values_only=True))
    
    file_count = 1
    row_count = 0
    csv_file = None
    writer = None
    
    for row in ws.iter_rows(values_only=True):
        # Start a new CSV when needed
        if row_count % max_rows_per_file == 0:
            if csv_file:
                csv_file.close()
            csv_filename = os.path.join(output_dir, f"part_{file_count}.csv")
            csv_file = open(csv_filename, "w", newline="", encoding="utf-8")
            writer = csv.writer(csv_file)
            writer.writerow(header)
            file_count += 1
        
        writer.writerow(row)
        row_count += 1
    
    if csv_file:
        csv_file.close()
    print(f"Split into {file_count-1} files in `{output_dir}`")

if __name__ == "__main__":
    split_excel_to_csvs(
        "/home/lp-55/Documents/Playground/angel-shot-python-data-migrate-script/Complete Brizo Database Snapshot Apr 7 2022 (1).xlsx",
        output_dir="./split_csvs",
        max_rows_per_file=50000
    )
